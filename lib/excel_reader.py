"""
excel_reader.py
AKOAKO_HACO_SNS投稿台帳.xlsx から投稿予定行を読み取るモジュール
"""

import re
from datetime import date, datetime
from pathlib import Path

import openpyxl

XLSX_PATH = Path(__file__).parent.parent / "AKOAKO_HACO_SNS投稿台帳.xlsx"

# シートごとの列マッピング (1-indexed)
# 実際のシート構造に合わせて調整してください
COLUMN_MAP = {
    "投稿日": 1,
    "テーマ": 2,
    "参照URL": 3,
    "画像URL": 4,
    "優先SNS": 5,
    "メモ": 6,
    "投稿済": 7,
    "生成日時": 8,
    # SNS別生成テキスト列 (後日拡張)
    "X文章": 9,
    "note文章": 10,
    "Threads文章": 11,
    "TikTok台本": 12,
}


def _parse_date(value) -> date | None:
    """
    投稿日セルの値を date に変換する。
    対応形式: date/datetime オブジェクト、
              文字列 "2026/05/12", "2026-05-12", "2026/5/12", "2026-5-12"
    """
    if value is None:
        return None
    if isinstance(value, (date, datetime)):
        return value.date() if isinstance(value, datetime) else value
    s = str(value).strip()
    # 数値文字列 (Excel シリアル値) は非対応、文字列パターンのみ
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y/%#m/%#d", "%Y-%-m-%-d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    # フォーマット文字列が環境依存のため正規表現でも試みる
    m = re.fullmatch(r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})", s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass
    return None


def _parse_sns_list(value) -> list[str]:
    """
    優先SNS セルを SNS名リストに変換する。
    例: "X, Threads" -> ["X", "Threads"]
    """
    if not value:
        return []
    return [s.strip() for s in re.split(r"[,、/]", str(value)) if s.strip()]


def _is_posted(ws, row_num: int) -> bool:
    """投稿済列に値が入っていれば True"""
    col = COLUMN_MAP["投稿済"]
    val = ws.cell(row=row_num, column=col).value
    if val is None:
        return False
    return str(val).strip() not in ("", "0", "False", "false")


def read_today_rows(sheet_name: str, target_date: date | None = None) -> list[dict]:
    """
    今日(または target_date)の投稿予定行を dict リストで返す。

    Parameters
    ----------
    sheet_name : str
        "AKOAKO" or "HACO_LABO"
    target_date : date, optional
        指定なしの場合は今日の日付を使用

    Returns
    -------
    list[dict]
        各行の情報を持つ dict のリスト。
        キー: row_num, 投稿日, テーマ, 参照URL, 画像URL, 優先SNS, メモ
    """
    if target_date is None:
        target_date = date.today()

    if not XLSX_PATH.exists():
        raise FileNotFoundError(f"台帳ファイルが見つかりません: {XLSX_PATH}")

    wb = openpyxl.load_workbook(str(XLSX_PATH), data_only=True)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"シート '{sheet_name}' が見つかりません。存在するシート: {wb.sheetnames}")

    ws = wb[sheet_name]
    results = []

    # 行1=ヘッダー、行2=凡例 → 行3以降をスキャン
    for row_num in range(3, ws.max_row + 1):
        posting_date = _parse_date(ws.cell(row=row_num, column=COLUMN_MAP["投稿日"]).value)
        if posting_date is None:
            continue
        if posting_date != target_date:
            continue
        if _is_posted(ws, row_num):
            continue

        row_data = {
            "row_num": row_num,
            "投稿日": posting_date,
            "テーマ": ws.cell(row=row_num, column=COLUMN_MAP["テーマ"]).value,
            "参照URL": ws.cell(row=row_num, column=COLUMN_MAP["参照URL"]).value,
            "画像URL": ws.cell(row=row_num, column=COLUMN_MAP["画像URL"]).value,
            "優先SNS": _parse_sns_list(ws.cell(row=row_num, column=COLUMN_MAP["優先SNS"]).value),
            "メモ": ws.cell(row=row_num, column=COLUMN_MAP["メモ"]).value,
        }
        results.append(row_data)

    wb.close()
    return results
